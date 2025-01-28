import os
import pandas as pd
from pathlib import Path
import io
import numpy as np
import re
import time
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
from typing import Optional
from concurrent.futures import ThreadPoolExecutor
import torch
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import streamlit as st
import pickle
from datetime import datetime
import glob

class TextAnalyzer:
    def __init__(self):
        try:
            # Initialize model
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
            st.info(f"Using device: {self.device}")
            
            try:
                self.model = SentenceTransformer('paraphrase-MPNet-base-v2', device=self.device)
                st.success("Successfully loaded primary model")
            except:
                st.warning("Attempting to use fallback model...")
                try:
                    self.model = SentenceTransformer('distilbert-base-nli-mean-tokens', device=self.device)
                    st.success("Successfully loaded fallback model")
                except:
                    st.error("No models could be loaded. The application cannot continue.")
                    st.stop()
        except Exception as e:
            st.error("Critical error during initialization")
            st.stop()

    def create_embeddings(self, texts, batch_size):
        try:
            text_map = {i: str(text) for i, text in enumerate(texts) if text and not pd.isna(text) and not str(text).isspace()}
            unique_texts = list(set(text_map.values()))
            
            embeddings = []
            total_batches = (len(unique_texts) + batch_size - 1) // batch_size
            
            progress_bar = st.progress(0)
            with ThreadPoolExecutor(max_workers=4) as executor:
                for batch_idx in range(0, len(unique_texts), batch_size):
                    try:
                        batch = unique_texts[batch_idx:batch_idx+batch_size]
                        batch_embeddings = self.model.encode(batch, show_progress_bar=False)
                        embeddings.extend(batch_embeddings)
                        
                        progress = min(1.0, (batch_idx + batch_size) / len(unique_texts))
                        progress_bar.progress(progress)
                        
                    except:
                        if not embeddings:  # If no embeddings created yet, fail
                            raise
                        break  # Otherwise use what we have
            
            progress_bar.empty()
            
            text_to_embedding = {text: emb for text, emb in zip(unique_texts, embeddings)}
            embedding_dim = embeddings[0].shape[0] if embeddings else self.model.get_sentence_embedding_dimension()
            
            final_embeddings = np.array([text_to_embedding[text_map[i]] if i in text_map 
                                       else np.zeros(embedding_dim) for i in range(len(texts))])
            
            return final_embeddings
            
        except Exception as e:
            st.error("Failed to create embeddings")
            st.stop()

    def calculate_cosine_similarity(self, embeddings, batch_size=1000):
        try:
            n = len(embeddings)
            cosine_sim = np.zeros((n, n))
            
            with st.spinner("Calculating cosine similarity..."):
                progress_bar = st.progress(0)
                total_batches = (n + batch_size - 1) // batch_size
                
                for batch_idx in range(0, n, batch_size):
                    try:
                        end = min(batch_idx + batch_size, n)
                        batch = embeddings[batch_idx:end]
                        cosine_sim[batch_idx:end] = cosine_similarity(batch, embeddings)
                        
                        progress = min(1.0, (batch_idx + batch_size) / n)
                        progress_bar.progress(progress)
                        
                        del batch  # Free memory
                        
                    except:
                        if batch_idx == 0:  # If first batch failed
                            raise
                        break  # Otherwise use partial results
                
                progress_bar.empty()
            
            return cosine_sim
            
        except Exception as e:
            st.error("Failed to calculate cosine similarity")
            st.stop()

    def process_excel(self, input_file: str, column: str) -> pd.DataFrame:
        try:
            df = pd.read_excel(input_file)
        except:
            st.error("Failed to read Excel file")
            st.stop()
            
        total_rows = len(df)
        st.write(f"Processing {total_rows} articles...")
        
        return df

    def save_state(self, state_data, state_name):
        """Save current state to a pickle file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"state_{state_name}_{timestamp}.pkl"
        try:
            with open(filename, 'wb') as f:
                pickle.dump(state_data, f)
            return filename
        except:
            st.error("Failed to save state")
            return None

    def load_state(self, filename):
        """Load state from a pickle file"""
        try:
            with open(filename, 'rb') as f:
                state_data = pickle.load(f)
            return state_data
        except:
            st.error("Failed to load state")
            return None

    def perform_topic_clustering(self, input_file: str, text_column: str, threshold1: float, threshold2: float, threshold3: float, batch_size: int = 500):
        try:
            df = pd.read_excel(input_file)
            if text_column not in df.columns:
                st.error(f"Column '{text_column}' not found in Excel file")
                return None, None

            # Check for existing embeddings
            embedding_file = "embeddings_state.pkl"
            if os.path.exists(embedding_file):
                if st.checkbox("Load existing embeddings?"):
                    embeddings = self.load_state(embedding_file)
                    if embeddings is not None:
                        st.success("Loaded existing embeddings")
                    else:
                        st.warning("Failed to load embeddings, computing new ones")
                        embeddings = None

            if 'embeddings' not in locals() or embeddings is None:
                with st.spinner("Creating embeddings..."):
                    texts = df[text_column].fillna('').tolist()
                    embeddings = self.create_embeddings(texts, batch_size)
                    # Save embeddings
                    self.save_state(embeddings, "embeddings")

            # Check for existing similarity matrix
            similarity_file = "similarity_state.pkl"
            if os.path.exists(similarity_file):
                if st.checkbox("Load existing similarity matrix?"):
                    cosine_sim = self.load_state(similarity_file)
                    if cosine_sim is not None:
                        st.success("Loaded existing similarity matrix")
                    else:
                        st.warning("Failed to load similarity matrix, computing new one")
                        cosine_sim = None

            if 'cosine_sim' not in locals() or cosine_sim is None:
                with st.spinner("Calculating cosine similarity..."):
                    cosine_sim = self.calculate_cosine_similarity(embeddings, batch_size=batch_size)
                    # Save similarity matrix
                    self.save_state(cosine_sim, "similarity")

            topic_data = {
                threshold1: {'topics': [], 'titles': [], 'counts': {}},
                threshold2: {'topics': [], 'titles': [], 'counts': {}},
                threshold3: {'topics': [], 'titles': [], 'counts': {}}
            }
            
            # Process for each threshold
            for threshold in [threshold1, threshold2, threshold3]:
                st.write(f"\nProcessing for similarity threshold: {threshold}")
                
                topics = [-1] * len(df)
                current_topic = 1
                
                empty_topic = 0
                for i in range(len(df)):
                    if pd.isna(df.iloc[i][text_column]) or str(df.iloc[i][text_column]).isspace():
                        topics[i] = empty_topic
                
                for i in range(len(df)):
                    if topics[i] == -1:
                        topics[i] = current_topic
                        break
                
                st.write("Assigning topics...")
                progress_bar = st.progress(0)
                for i in range(1, len(df)):
                    progress_bar.progress(i / len(df))
                    if topics[i] != -1:
                        continue
                        
                    found_similar = False
                    for j in range(i):
                        if topics[j] > 0 and cosine_sim[i][j] >= threshold:
                            topics[i] = topics[j]
                            found_similar = True
                            break
                    
                    if not found_similar:
                        current_topic += 1
                        topics[i] = current_topic
                
                progress_bar.empty()
                
                df[f'Topics_threshold_{threshold}'] = [f"Empty Content" if x == 0 else f"Topic {x}" if x != -1 else "Unassigned" for x in topics]
                
                st.write("Generating topic titles...")
                df[f'Topic_Title_threshold_{threshold}'] = ''
                topic_titles = {}
                topic_counts = {}
                
                topic_titles["Empty Content"] = "Empty Content"
                topic_counts["Empty Content"] = topics.count(0)
                
                progress_bar = st.progress(0)
                for topic_num in range(1, current_topic + 1):
                    progress_bar.progress(topic_num / current_topic)
                    try:
                        topic_name = f"Topic {topic_num}"
                        topic_texts = df[df[f'Topics_threshold_{threshold}'] == topic_name][text_column]
                        
                        topic_counts[topic_name] = len(topic_texts)
                        
                        if not topic_texts.empty:
                            topic_embeddings = self.model.encode(topic_texts.tolist(), show_progress_bar=False)
                            avg_embedding = np.mean(topic_embeddings, axis=0)
                            similarities = cosine_similarity([avg_embedding], topic_embeddings)[0]
                            top_index = similarities.argmax()
                            
                            full_text = topic_texts.iloc[top_index]
                            abbreviations = [
                                "Pte. Ltd.",
                                "Co.",
                                "Inc.",
                                "Ltd.",
                                "Corp."
                            ]
                            
                            protected_text = re.sub(r'\b([A-Z]\.)\s*', lambda m: f"INITIAL{ord(m.group(1)[0])}", str(full_text))
                            
                            for i, abbr in enumerate(abbreviations):
                                protected_text = protected_text.replace(abbr, f"ABBR{i}")
                            
                            sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', protected_text)
                            
                            for i, abbr in enumerate(abbreviations):
                                sentences = [s.replace(f"ABBR{i}", abbr) for s in sentences]
                            
                            sentences = [re.sub(r'INITIAL(\d+)', lambda m: chr(int(m.group(1))) + ".", s) for s in sentences]
                            
                            topic_title = sentences[0] if sentences else str(full_text)
                            topic_titles[topic_name] = topic_title
                            
                    except:
                        topic_titles[topic_name] = f"Untitled Topic {topic_num}"
                
                progress_bar.empty()
                
                topic_data[threshold]['topics'] = list(range(1, current_topic + 1))
                topic_data[threshold]['titles'] = topic_titles
                topic_data[threshold]['counts'] = topic_counts
                
                for topic_name, title in topic_titles.items():
                    df.loc[df[f'Topics_threshold_{threshold}'] == topic_name, f'Topic_Title_threshold_{threshold}'] = title
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Results', index=False)
                
                sheet_name = 'Topics_All_Thresholds'
                pivot_df = pd.DataFrame(columns=['Topic', 'Topic Title', 'Document Count', 'Titles in Topic'])
                
                current_row = 0
                for threshold in [threshold1, threshold2, threshold3]:
                    pivot_df.loc[current_row] = [f'Threshold {threshold}', '', '', '']
                    current_row += 1
                    
                    for topic_name in sorted(df[f'Topics_threshold_{threshold}'].unique()):
                        topic_docs = df[df[f'Topics_threshold_{threshold}'] == topic_name]
                        topic_title = topic_docs[f'Topic_Title_threshold_{threshold}'].iloc[0]
                        titles_in_topic = topic_docs['Title'].tolist() if 'Title' in df.columns else []
                        
                        pivot_df.loc[current_row] = [topic_name, topic_title, len(topic_docs), '']
                        current_row += 1
                        
                        for title in titles_in_topic:
                            pivot_df.loc[current_row] = ['', '', '', title]
                            current_row += 1
                    
                    pivot_df.loc[current_row] = ['', '', '', '']
                    current_row += 1
                
                pivot_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                try:
                    worksheet = writer.sheets[sheet_name]
                    
                    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    threshold_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
                    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
                    header_font = Font(bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    for row in worksheet.iter_rows(min_row=2):
                        if row[0].value and str(row[0].value).startswith('Threshold'):
                            worksheet.merge_cells(
                                start_row=row[0].row,
                                start_column=1,
                                end_row=row[0].row,
                                end_column=4
                            )
                            row[0].fill = threshold_fill
                            row[0].font = header_font
                            row[0].alignment = Alignment(horizontal='center', vertical='center')
                            for cell in row[:4]:
                                cell.border = border
                    
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                        for cell in row:
                            cell.border = border
                            
                            if cell.column == 1:
                                cell.alignment = Alignment(horizontal='center')
                            elif cell.column == 2:
                                cell.alignment = Alignment(horizontal='left')
                            elif cell.column == 3:
                                cell.alignment = Alignment(horizontal='center')
                            elif cell.column == 4:
                                cell.alignment = Alignment(horizontal='left')
                                if cell.value:
                                    worksheet.row_dimensions[row_idx].outline_level = 1
                    
                    worksheet.column_dimensions['A'].width = 15
                    worksheet.column_dimensions['B'].width = 60
                    worksheet.column_dimensions['C'].width = 15
                    worksheet.column_dimensions['D'].width = 60
                    
                    worksheet.sheet_properties.outlinePr.summaryBelow = False
                    for row in worksheet.row_dimensions:
                        if worksheet.row_dimensions[row].outline_level == 1:
                            worksheet.row_dimensions[row].hidden = True
                            
                except:
                    st.warning("Error formatting Excel worksheet")
                    # Continue without formatting
            
            output.seek(0)
            st.success("Successfully completed topic clustering")
            return df, output
            
        except Exception as e:
            st.error("Error processing file")
            return None, None

def main():
    try:
        st.title("Topic Clustering Tool")
        
        # Add state management section 
        with st.expander("State Management"):
            st.write("Load or clear saved states")
            if st.button("Clear All Saved States"):
                try:
                    for f in glob.glob("state_*.pkl"):
                        os.remove(f)
                    st.success("Cleared all saved states")
                except:
                    st.error("Error clearing states")

        try:
            analyzer = TextAnalyzer()
        except:
            st.error("Failed to initialize TextAnalyzer")
            st.stop()
            
        uploaded_file = st.file_uploader("Upload your Excel file", type=['xlsx'])
        
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file)
                columns = df.columns.tolist()
                
                with st.form("clustering_parameters"):
                    text_column = st.selectbox("Select column for topic clustering", columns)
                    
                    st.info("Similarity Score Guide: A lower threshold (closer to 0) will create more general topics with broader groupings, while a higher threshold (closer to 1) will create more specific, tightly focused topics.")
                    
                    threshold1 = st.slider("First similarity threshold", 0.0, 1.0, 0.7, 0.01)
                    threshold2 = st.slider("Second similarity threshold", 0.0, 1.0, 0.8, 0.01)
                    threshold3 = st.slider("Third similarity threshold", 0.0, 1.0, 0.9, 0.01)
                    batch_size = st.number_input("Batch size", 100, 1000, 500)
                    
                    submit_button = st.form_submit_button("Start Topic Clustering")
                
                if submit_button:
                    try:
                        with st.spinner("Processing topic clustering..."):
                            clustering_results, output = analyzer.perform_topic_clustering(
                                uploaded_file, text_column,
                                threshold1, threshold2, threshold3, batch_size
                            )
                            
                            if clustering_results is not None and output is not None:
                                st.success("Topic clustering completed successfully")
                                
                                st.download_button(
                                    label="Download Topic Clustering Results",
                                    data=output.getvalue(),
                                    file_name="topic_clustering_results.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.error("Topic clustering produced no results")
                    except Exception as e:
                        st.error(f"Error during topic clustering: {str(e)}")
                        
            except:
                st.error("Error reading Excel file")
                
    except:
        st.error("Critical error in main application")

if __name__ == "__main__":
    main()
